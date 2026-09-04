import os
from openpyxl import load_workbook
from datetime import datetime
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak, Image
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

class FlightCrewDeclaration:
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.flights_data = []
        self.load_data()
    
    def load_data(self):
        """Load flight data from Excel file"""
        try:
            wb = load_workbook(self.excel_file)
            ws = wb.active
            
            # Parse flight schedule data
            current_date = None
            for row in ws.iter_rows(min_row=5, values_only=True):
                if row[0] is None:
                    continue
                
                # Extract flight info
                if row[0] and str(row[0]).startswith('09/'):  # DATE column
                    current_date = row[0]
                    flight_number = row[1]  # FLT
                    aircraft_type = row[2]  # TYPE
                    aircraft_reg = row[3]  # REG
                    departure = row[4]  # DEP
                    arrival = row[4]  # ARR (same as DEP in this case, adjust if needed)
                    std = row[5]  # Standard Departure Time
                    eta = row[7]  # Estimated Time of Arrival
                    crew_count = row[10]  # Crew #
                    crew_list = row[11]  # Crew names
                    
                    if flight_number:
                        self.flights_data.append({
                            'date': current_date,
                            'flight': flight_number,
                            'aircraft_type': aircraft_type,
                            'aircraft_reg': aircraft_reg,
                            'departure': departure,
                            'arrival': arrival,
                            'std': std,
                            'eta': eta,
                            'crew_count': crew_count,
                            'crew_list': crew_list
                        })
            
            wb.close()
            print(f"✅ Loaded {len(self.flights_data)} flights from Excel")
        
        except Exception as e:
            print(f"❌ Error loading Excel: {e}")
            return False
        
        return True
    
    def parse_crew_list(self, crew_text):
        """Parse crew names from the crew column"""
        if not crew_text:
            return []
        
        crew_members = []
        lines = str(crew_text).split('\n')
        
        for line in lines:
            line = line.strip()
            if line and not line.startswith('-'):
                crew_members.append(line)
            elif line.startswith('-'):
                crew_members.append(line[1:].strip())
        
        return crew_members
    
    def generate_pdf(self, output_file='General_Declaration.pdf'):
        """Generate PDF for all flights"""
        if not self.flights_data:
            print("❌ No flight data to generate PDF")
            return False
        
        try:
            doc = SimpleDocTemplate(output_file, pagesize=A4,
                                  topMargin=0.5*inch, bottomMargin=0.5*inch,
                                  leftMargin=0.5*inch, rightMargin=0.5*inch)
            
            story = []
            styles = getSampleStyleSheet()
            
            # Create custom styles
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                textColor=colors.black,
                spaceAfter=6,
                alignment=TA_CENTER,
                fontName='Helvetica-Bold'
            )
            
            heading_style = ParagraphStyle(
                'CustomHeading',
                parent=styles['Heading2'],
                fontSize=10,
                textColor=colors.black,
                spaceAfter=6,
                fontName='Helvetica-Bold'
            )
            
            normal_style = ParagraphStyle(
                'CustomNormal',
                parent=styles['Normal'],
                fontSize=9,
                spaceAfter=3,
                fontName='Helvetica'
            )
            
            # Process each flight
            for idx, flight in enumerate(self.flights_data):
                if idx > 0:
                    story.append(PageBreak())
                
                # Header
                story.append(Paragraph("GENERAL DECLARATION", title_style))
                story.append(Paragraph("(OUTWARD/INWARD)", title_style))
                story.append(Spacer(1, 0.2*inch))
                
                # Flight Information Table
                flight_info = [
                    ['Operator', 'Sun Phu Quoc Airways', 'Aircraft', '32E'],
                    ['Marks of Nationality and Registration', flight['aircraft_reg'], 'Flight', flight['flight']],
                    ['Departure from', self._airport_code(flight['departure']), 'Date', self._format_date(flight['date'])],
                    ['', '', 'Arrival at', self._airport_code(flight['arrival'])]
                ]
                
                flight_table = Table(flight_info, colWidths=[1.8*inch, 1.8*inch, 1.8*inch, 1.8*inch])
                flight_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, -1), colors.lightgrey),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 8),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                    ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(flight_table)
                story.append(Spacer(1, 0.15*inch))
                
                # Crew List
                story.append(Paragraph("FLIGHT ROUTING", heading_style))
                
                crew_members = self.parse_crew_list(flight['crew_list'])
                crew_data = [['Port', 'ID', 'Name', 'Pax', 'Pos', 'Passport', 'Birth Date', 'G', 'Ntly', 'Passport', 'No. Of Passengers']]
                
                # Add crew members
                port_from = self._airport_code(flight['departure'])
                port_to = self._airport_code(flight['arrival'])
                
                for idx_crew, member in enumerate(crew_members[:10]):  # Limit to 10 crew members
                    if idx_crew == 0:
                        crew_data.append([port_from, '', member, '', '', '', '', '', 'VNM', '', ''])
                    else:
                        crew_data.append(['', '', member, '', '', '', '', '', 'VNM', '', ''])
                
                # Add arrival port
                crew_data.append([port_to, '', '', '', '', '', '', '', '', '', ''])
                
                crew_table = Table(crew_data, colWidths=[0.6*inch, 0.6*inch, 1.5*inch, 0.4*inch, 0.4*inch, 0.8*inch, 0.7*inch, 0.3*inch, 0.4*inch, 0.8*inch, 1.2*inch])
                crew_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, -1), 7),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                    ('TOPPADDING', (0, 0), (-1, -1), 3),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.black)
                ]))
                story.append(crew_table)
                story.append(Spacer(1, 0.2*inch))
                
                # Declaration footer
                story.append(Paragraph("DECLARATION OF HEALTH", heading_style))
                story.append(Spacer(1, 0.3*inch))
                story.append(Paragraph("Signature: ___________________________________________________________", normal_style))
                story.append(Spacer(1, 0.1*inch))
                story.append(Paragraph("Authorized Agent or Pilot-In-Command", normal_style))
            
            # Build PDF
            doc.build(story)
            print(f"✅ PDF generated successfully: {output_file}")
            return True
        
        except Exception as e:
            print(f"❌ Error generating PDF: {e}")
            return False
    
    def _airport_code(self, airport):
        """Convert airport codes"""
        airport_map = {
            'HAN': 'HAN (HANOI)',
            'SGN': 'SGN (SAIGON)',
            'DAD': 'DAD (DANANG)',
            'PQC': 'PQC (PHU QUOC)',
            'HKG': 'HKG (HONG KONG)'
        }
        return airport_map.get(str(airport), str(airport))
    
    def _format_date(self, date_obj):
        """Format date for display"""
        if isinstance(date_obj, str):
            return date_obj
        if hasattr(date_obj, 'strftime'):
            return date_obj.strftime('%d/%m/%Y')
        return str(date_obj)


def main():
    print("=" * 60)
    print("FLIGHT CREW DECLARATION GENERATOR")
    print("=" * 60)
    
    # Get input file
    excel_file = input("\n📁 Enter Excel file path (or press Enter for 'flight_schedule.xlsx'): ").strip()
    if not excel_file:
        excel_file = 'flight_schedule.xlsx'
    
    # Check if file exists
    if not os.path.exists(excel_file):
        print(f"❌ File not found: {excel_file}")
        return
    
    # Get output file
    output_file = input("📄 Enter PDF output filename (or press Enter for 'General_Declaration.pdf'): ").strip()
    if not output_file:
        output_file = 'General_Declaration.pdf'
    
    if not output_file.endswith('.pdf'):
        output_file += '.pdf'
    
    # Generate declaration
    print("\n⏳ Processing flight data...")
    declaration = FlightCrewDeclaration(excel_file)
    
    if declaration.generate_pdf(output_file):
        print(f"\n✅ Success! PDF saved as: {output_file}")
    else:
        print("\n❌ Failed to generate PDF")


if __name__ == "__main__":
    main()
